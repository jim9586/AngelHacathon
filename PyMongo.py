__author__ = 'lizhengning1'


import datetime
import openpyxl
import xlrd as xl
import requests
from flask import Flask, jsonify

app = Flask(__name__)


states = {
        'AK': 'Alaska',
        'AL': 'Alabama',
        'AR': 'Arkansas',
        'AZ': 'Arizona',
        'CA': 'California',
        'CO': 'Colorado',
        'CT': 'Connecticut',
        'DC': 'District of Columbia',
        'DE': 'Delaware',
        'FL': 'Florida',
        'GA': 'Georgia',
        'HI': 'Hawaii',
        'IA': 'Iowa',
        'ID': 'Idaho',
        'IL': 'Illinois',
        'IN': 'Indiana',
        'KS': 'Kansas',
        'KY': 'Kentucky',
        'LA': 'Louisiana',
        'MA': 'Massachusetts',
        'MD': 'Maryland',
        'ME': 'Maine',
        'MI': 'Michigan',
        'MN': 'Minnesota',
        'MO': 'Missouri',
        'MS': 'Mississippi',
        'MT': 'Montana',
        'NC': 'North Carolina',
        'ND': 'North Dakota',
        'NE': 'Nebraska',
        'NH': 'New Hampshire',
        'NJ': 'New Jersey',
        'NM': 'New Mexico',
        'NV': 'Nevada',
        'NY': 'New York',
        'OH': 'Ohio',
        'OK': 'Oklahoma',
        'OR': 'Oregon',
        'PA': 'Pennsylvania',
        'RI': 'Rhode Island',
        'SC': 'South Carolina',
        'SD': 'South Dakota',
        'TN': 'Tennessee',
        'TX': 'Texas',
        'UT': 'Utah',
        'VA': 'Virginia',
        'VT': 'Vermont',
        'WA': 'Washington',
        'WI': 'Wisconsin',
        'WV': 'West Virginia',
        'WY': 'Wyoming'
}


# for i in range(8, 60):
#     data = {
#         "states":"",
#         "VCDeals":[]
#     }
#     data['states'] = sheet['A' + str(i)].value
#     for k in range(0, 6):
#         tmp = {
#         "Deals":"",
#         "Amount":""
#         }
#         tmp['Deals'] = sheet[deallist[k] + str(i)].value
#         tmp['Amount'] = sheet[amountlist[k] + str(i)].value
#         tmplist.append(tmp)
#     data['VCDeals'] = tmplist
#     datalist.append(data)
#     tmplist = []


def get_data():
    vxls = openpyxl.load_workbook('VC.xlsx')
    wb = openpyxl.load_workbook('Job.xlsx')
    xls = xl.open_workbook('Job.xlsx')
    Firm_wb = openpyxl.load_workbook('Firm.xlsx')
    Firm_xls = xl.open_workbook('Firm.xlsx')
    tmplist = []
    deallist = ['B', 'D', 'F', 'H', 'J', 'L']
    amountlist = ['C','E','G','I','K','M']
    datalist = []
    empList = []
    joblist = []
    finalList = []
    firmList = []
    sheetVC = vxls.get_sheet_by_name('Sheet1')
    m = 0

    for key in sorted(states.keys()):
        score1 = 0
        score2 = 0
        score3 = 0
        data = {"GDP": "",
                "states": "",
                "UnEmployRate": "",
                "JobGrowthRate": "",
                "VCDeals": "",
                "Score": "",
                "Firm": ""}
        result = requests.get('https://www.quandl.com/api/v3/datasets/FRED/' + key +
                          'NGSP.json?api_key=FdR_cvbf6vXrTPcYGVR4&start_date=2005-01-01&end_date=2014-12-31')
        for Gkey in result.json()['dataset']['data']:
            for i in range(0,10):
                tmp = result.json()['dataset']['data'][i][1]
                datalist.append(tmp)
                data["GDP"] = datalist
                score1 += 0.00008 * tmp
                data["states"] = key
            datalist = []
        if key in xls.sheet_names():

            sheet = wb.get_sheet_by_name(key)
            i = 5
            while i < 125:
                # print sheet['A' + str(i)].value
                original = sheet['B' + str(i)].value
                current = sheet['B' + str(i + 12)].value
                increase = float(current - original) / original
                joblist.append(increase)
                score2 += (increase * 0.5) / 100
                data['JobGrowthRate'] = joblist
                i = i + 12
            joblist = []
        else:
            data['JobGrowthRate'] = None
        if key in Firm_xls.sheet_names():
            sheet_Firm = Firm_wb.get_sheet_by_name(key)
            for count in range(39, 48):
                firmList.append(sheet_Firm['B' + str(count)].value)
            data['Firm'] = firmList
            firmList = []
        else:
            data['Firm'] = None
        for k in range(0, 6):
            tmp = {
                "Deals":"",
                "Amount":""
                }
            #print deallist[k] + str(j + 8)
            tmp['Deals'] = sheetVC[deallist[k] + str(m + 8)].value
            tmp['Amount'] = sheetVC[amountlist[k] + str(m + 8)].value
            tmplist.append(tmp)
        data['VCDeals'] = tmplist
        m += 1
        tmplist = []
        resultUnEmploy = requests.get('https://www.quandl.com/api/v3/datasets/FRED/'
                                        + key +'UR.json?auth_token=FdR_cvbf6vXrTPcYGVR4&start_date=2005-01-01&end_date=2014-12-31')
        for newkey in resultUnEmploy.json()['dataset']['data']:
            for j in range(0,10):
                tmp = resultUnEmploy.json()['dataset']['data'][j][1]
                empList.append(tmp)
                score3 += (tmp * 0.5) / 100
                data["UnEmployRate"] = empList
            empList = []
        data['Score'] = (score2 + score3) * 2
        finalList.append(data)

    return finalList

data = get_data()

    # client = MongoClient()
    # # client = MongoClient('localhost', 27017)
    # db = client.test
    #
    # db = client.test_database
    # db = client['final_database']
    #
    # posts = db.posts
    # post_id = posts.insert_one(data).inserted_id
    # db.collection_names(include_system_collections=False)
    # print posts.find_one({"_id": post_id})
@app.route('/free/api/v1.0/tasks', methods=['GET'])
def get_tasks():
    return jsonify({'tasks': data})


if __name__ == '__main__':
    app.run(host='0.0.0.0')



